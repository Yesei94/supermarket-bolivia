from fastapi import FastAPI, Depends, HTTPException, Header
from fastapi.responses import StreamingResponse
from fastapi.middleware.cors import CORSMiddleware

from pydantic import BaseModel
from sqlalchemy import create_engine, Column, Integer, Float, String, DateTime
from sqlalchemy.orm import declarative_base, sessionmaker
from datetime import datetime, date
from io import BytesIO
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from auth import verify_token
import httpx, os
import openpyxl

app = FastAPI(title="Sales Service", version="1.0.0")
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

engine = create_engine("sqlite:///sales.db", connect_args={"check_same_thread": False})
SessionLocal = sessionmaker(bind=engine)
Base = declarative_base()

PRODUCT_URL = os.getenv("PRODUCT_URL", "http://product-service:8002")
INVENTORY_URL = os.getenv("INVENTORY_URL", "http://inventory-service:8003")
CUSTOMER_URL = os.getenv("CUSTOMER_URL", "http://customer-service:8004")
NOTIFICATION_URL = os.getenv("NOTIFICATION_URL", "http://notification-service:8005")

class Sale(Base):
    __tablename__ = "sales"
    id = Column(Integer, primary_key=True)
    customer_id = Column(Integer)
    customer_name = Column(String)
    product_id = Column(Integer)
    product_name = Column(String)
    branch_id = Column(Integer)
    quantity = Column(Integer)
    unit_price = Column(Float)
    unit_cost = Column(Float)
    total = Column(Float)
    profit = Column(Float)
    payment_type = Column(String)
    invoice_number = Column(String)
    date = Column(DateTime, default=datetime.utcnow)

Base.metadata.create_all(bind=engine)

class SaleRequest(BaseModel):
    customer_id: int
    product_id: int
    branch_id: int
    quantity: int
    payment_type: str = "Efectivo"

@app.get("/health")
def health():
    return {"status": "ok", "service": "sales-service"}

@app.post("/sales")
async def create_sale(request: SaleRequest, authorization: str = Header(None), token=Depends(verify_token)):
    headers = {"Authorization": authorization}
    async with httpx.AsyncClient() as client:
        product_resp = await client.get(f"{PRODUCT_URL}/products/{request.product_id}", headers=headers)
        if product_resp.status_code != 200: raise HTTPException(status_code=400, detail="Producto no encontrado")
        product = product_resp.json()

        customer_resp = await client.get(f"{CUSTOMER_URL}/customers/{request.customer_id}", headers=headers)
        if customer_resp.status_code != 200: raise HTTPException(status_code=400, detail="Cliente no encontrado")
        customer = customer_resp.json()

        inv_resp = await client.get(f"{INVENTORY_URL}/inventory/product/{request.product_id}", headers=headers)
        if inv_resp.status_code != 200: raise HTTPException(status_code=400, detail="No se pudo consultar inventario")
        inventory_list = inv_resp.json()
        branch_inv = next((i for i in inventory_list if i["branch_id"] == request.branch_id), None)
        if not branch_inv: raise HTTPException(status_code=400, detail="No existe stock en la sucursal")
        if branch_inv["stock_actual"] < request.quantity: raise HTTPException(status_code=400, detail="Stock insuficiente")

        unit_price = float(product["base_price"])
        unit_cost = float(branch_inv["cost"])
        total = round(unit_price * request.quantity, 2)
        profit = round((unit_price - unit_cost) * request.quantity, 2)
        invoice = f"FAC-{datetime.now().strftime('%Y%m%d')}-{int(datetime.now().timestamp())}"

        out_resp = await client.post(f"{INVENTORY_URL}/inventory/output", headers=headers, json={"product_id": request.product_id, "branch_id": request.branch_id, "quantity": request.quantity, "reason": "Venta"})
        if out_resp.status_code != 200: raise HTTPException(status_code=400, detail="No se pudo descontar inventario")

        points = int(total // 10)
        await client.post(f"{CUSTOMER_URL}/customers/{request.customer_id}/points", headers=headers, json={"points": points})
        await client.post(f"{NOTIFICATION_URL}/notifications", headers=headers, json={"customer_id": request.customer_id, "type": "WhatsApp", "content": f"Venta completada. Factura {invoice}. Total {total} Bs"})

    db = SessionLocal()
    sale = Sale(customer_id=request.customer_id, customer_name=customer["name"], product_id=request.product_id, product_name=product["name"], branch_id=request.branch_id, quantity=request.quantity, unit_price=unit_price, unit_cost=unit_cost, total=total, profit=profit, payment_type=request.payment_type, invoice_number=invoice)
    db.add(sale); db.commit(); db.refresh(sale); db.close()
    return {"event": "SaleCompleted", "invoice": invoice, "sale": sale}

@app.get("/sales")
def list_sales(token=Depends(verify_token)):
    db = SessionLocal(); data = db.query(Sale).all(); db.close(); return data

@app.get("/reports/sales/today")
def sales_today(token=Depends(verify_token)):
    db = SessionLocal(); today = date.today(); data = db.query(Sale).all()
    filtered = [s for s in data if s.date.date() == today]
    total = round(sum(s.total for s in filtered), 2)
    profit = round(sum(s.profit for s in filtered), 2)
    by_payment = {}
    for s in filtered:
        by_payment[s.payment_type] = round(by_payment.get(s.payment_type, 0) + s.total, 2)
    db.close()
    return {"date": str(today), "totalIncome": total, "totalProfit": profit, "byPaymentType": by_payment, "sales": filtered}

@app.get("/accounting/profit")
def profit(token=Depends(verify_token)):
    db = SessionLocal(); data = db.query(Sale).all()
    total_income = round(sum(s.total for s in data), 2)
    total_cost = round(sum(s.unit_cost * s.quantity for s in data), 2)
    total_profit = round(sum(s.profit for s in data), 2)
    db.close()
    return {"totalIncome": total_income, "totalCost": total_cost, "profit": total_profit}

@app.get("/reports/sales.xlsx")
def export_sales_excel(token=Depends(verify_token)):
    db = SessionLocal()
    sales = db.query(Sale).all()
    db.close()

    wb = openpyxl.Workbook()

    ws_summary = wb.active
    ws_summary.title = "RESUMEN"
    ws_summary.sheet_view.showGridLines = False
    ws_summary.freeze_panes = "A4"
    total_income = round(sum(s.total for s in sales), 2)
    total_cost = round(sum(s.unit_cost * s.quantity for s in sales), 2)
    total_profit = round(sum(s.profit for s in sales), 2)
    title_fill = PatternFill("solid", fgColor="0B1F3A")
    header_fill = PatternFill("solid", fgColor="1D4ED8")
    metric_fill = PatternFill("solid", fgColor="EFF6FF")
    title_font = Font(color="FFFFFF", bold=True, size=14)
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws_summary.merge_cells("A1:B1")
    ws_summary["A1"] = "RESUMEN PROFESIONAL DE VENTAS"
    ws_summary["A1"].fill = title_fill
    ws_summary["A1"].font = title_font
    ws_summary["A1"].alignment = Alignment(horizontal="center", vertical="center")

    ws_summary.append(["MÉTRICA", "VALOR"])
    for cell in ws_summary[2]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal="center")

    summary_rows = [
        ["INGRESOS TOTALES", total_income],
        ["COSTO TOTAL", total_cost],
        ["GANANCIA TOTAL", total_profit],
        ["CANTIDAD DE VENTAS", len(sales)],
    ]
    for row in summary_rows:
        ws_summary.append(row)
        current_row = ws_summary.max_row
        for cell in ws_summary[current_row]:
            cell.border = border
            cell.fill = metric_fill
        ws_summary[f"B{current_row}"].alignment = Alignment(horizontal="right")
        if current_row in (3, 4, 5):
            ws_summary[f"B{current_row}"].number_format = '#,##0.00'
        if current_row == 6:
            ws_summary[f"B{current_row}"].number_format = '#,##0'

    ws_summary.column_dimensions["A"].width = 26
    ws_summary.column_dimensions["B"].width = 18

    ws_detail = wb.create_sheet(title="VENTAS")
    ws_detail.sheet_view.showGridLines = False
    ws_detail.freeze_panes = "A3"
    ws_detail.auto_filter.ref = "A2:J2"
    detail_title_fill = PatternFill("solid", fgColor="0B1F3A")
    detail_header_fill = PatternFill("solid", fgColor="1D4ED8")
    detail_fill = PatternFill("solid", fgColor="F8FAFC")
    ws_detail.merge_cells("A1:J1")
    ws_detail["A1"] = "DETALLE DE VENTAS"
    ws_detail["A1"].fill = detail_title_fill
    ws_detail["A1"].font = title_font
    ws_detail["A1"].alignment = Alignment(horizontal="center", vertical="center")

    ws_detail.append(["FECHA", "FACTURA", "CLIENTE", "PRODUCTO", "SUCURSAL", "CANTIDAD", "PRECIO UNITARIO", "TOTAL", "GANANCIA", "FORMA DE PAGO"])
    for cell in ws_detail[2]:
        cell.fill = detail_header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal="center")
    for sale in sales:
        ws_detail.append([
            sale.date.strftime("%Y-%m-%d %H:%M:%S") if sale.date else "",
            sale.invoice_number,
            sale.customer_name,
            sale.product_name,
            sale.branch_id,
            sale.quantity,
            sale.unit_price,
            sale.total,
            sale.profit,
            sale.payment_type,
        ])
        current_row = ws_detail.max_row
        for cell in ws_detail[current_row]:
            cell.border = border
            cell.fill = detail_fill
            cell.alignment = Alignment(vertical="center")
        for col in (6, 7, 8, 9):
            ws_detail.cell(current_row, col).number_format = '#,##0.00'

    totals_row = ws_detail.max_row + 2
    ws_detail[f"A{totals_row}"] = "TOTAL"
    ws_detail[f"F{totals_row}"] = sum(s.quantity for s in sales)
    ws_detail[f"H{totals_row}"] = total_income
    ws_detail[f"I{totals_row}"] = total_profit
    for ref in (f"A{totals_row}", f"F{totals_row}", f"H{totals_row}", f"I{totals_row}"):
        ws_detail[ref].font = Font(bold=True)
    ws_detail[f"F{totals_row}"].number_format = '#,##0'
    ws_detail[f"H{totals_row}"].number_format = '#,##0.00'
    ws_detail[f"I{totals_row}"].number_format = '#,##0.00'

    for column, width in {"A": 20, "B": 24, "C": 26, "D": 28, "E": 20, "F": 12, "G": 16, "H": 14, "I": 14, "J": 16}.items():
        ws_detail.column_dimensions[column].width = width

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    headers = {"Content-Disposition": 'attachment; filename="REPORTE_VENTAS.xlsx"'}
    return StreamingResponse(buffer, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers)
