from fastapi import FastAPI, Depends, HTTPException, UploadFile, File, Header
from fastapi.responses import StreamingResponse
from fastapi.middleware.cors import CORSMiddleware

from pydantic import BaseModel
from sqlalchemy import create_engine, Column, Integer, String, Float, DateTime
from sqlalchemy.orm import declarative_base, sessionmaker
from datetime import datetime
from io import BytesIO
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
import httpx
from auth import verify_token
import openpyxl
import os

app = FastAPI(title="Inventory Service", version="1.0.0")
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

engine = create_engine("sqlite:///inventory.db", connect_args={"check_same_thread": False})
SessionLocal = sessionmaker(bind=engine)
Base = declarative_base()
NOTIFICATION_URL = os.getenv("NOTIFICATION_URL", "http://notification-service:8005")

class Inventory(Base):
    __tablename__ = "inventory"
    id = Column(Integer, primary_key=True)
    product_id = Column(Integer)
    product_code = Column(String)
    product_name = Column(String)
    company_name = Column(String)
    branch_id = Column(Integer)
    branch_name = Column(String)
    stock_actual = Column(Integer, default=0)
    stock_minimo = Column(Integer, default=10)
    cost = Column(Float, default=0)
    price = Column(Float, default=0)

class Kardex(Base):
    __tablename__ = "kardex"
    id = Column(Integer, primary_key=True)
    product_id = Column(Integer)
    product_name = Column(String)
    branch_id = Column(Integer)
    branch_name = Column(String)
    movement_type = Column(String)
    quantity = Column(Integer)
    stock_final = Column(Integer)
    detail = Column(String)
    date = Column(DateTime, default=datetime.utcnow)

class EventLog(Base):
    __tablename__ = "event_logs"
    id = Column(Integer, primary_key=True)
    event_name = Column(String)
    content = Column(String)
    date = Column(DateTime, default=datetime.utcnow)

Base.metadata.create_all(bind=engine)

class InputRequest(BaseModel):
    product_id: int
    product_code: str
    product_name: str
    company_name: str = "OXXO Bolivia"
    branch_id: int
    branch_name: str
    quantity: int
    cost: float
    price: float
    stock_minimo: int = 10

class OutputRequest(BaseModel):
    product_id: int
    branch_id: int
    quantity: int
    reason: str = "Salida"

class TransferRequest(BaseModel):
    product_id: int
    branch_origin_id: int
    branch_destination_id: int
    branch_destination_name: str = "Sucursal destino"
    quantity: int

def log(db, name, content):
    db.add(EventLog(event_name=name, content=content)); db.commit()

def find_inv(db, product_id, branch_id):
    return db.query(Inventory).filter(Inventory.product_id == product_id, Inventory.branch_id == branch_id).first()

def add_kardex(db, inv, typ, qty, detail):
    db.add(Kardex(product_id=inv.product_id, product_name=inv.product_name, branch_id=inv.branch_id, branch_name=inv.branch_name, movement_type=typ, quantity=qty, stock_final=inv.stock_actual, detail=detail))

async def send_notification(authorization, content):
    headers = {"Authorization": authorization} if authorization else {}
    async with httpx.AsyncClient() as client:
        try:
            await client.post(f"{NOTIFICATION_URL}/notifications", headers=headers, json={"customer_id": 0, "type": "STOCK BAJO", "content": content})
        except Exception:
            pass

@app.get("/health")
def health():
    return {"status": "ok", "service": "inventory-service"}

@app.post("/inventory/input")
async def input_inventory(request: InputRequest, authorization: str = Header(None), token=Depends(verify_token)):
    db = SessionLocal()
    inv = find_inv(db, request.product_id, request.branch_id)
    if not inv:
        inv = Inventory(product_id=request.product_id, product_code=request.product_code, product_name=request.product_name, company_name=request.company_name, branch_id=request.branch_id, branch_name=request.branch_name, stock_actual=0, stock_minimo=request.stock_minimo, cost=request.cost, price=request.price)
        db.add(inv); db.commit(); db.refresh(inv)
    inv.stock_actual += request.quantity
    inv.cost = request.cost; inv.price = request.price; inv.stock_minimo = request.stock_minimo
    add_kardex(db, inv, "ENTRADA", request.quantity, "Ingreso de mercadería")
    db.commit(); db.refresh(inv)
    log(db, "InventoryUpdated", f"Ingreso {request.quantity} de {request.product_name} en {request.branch_name}")
    if inv.stock_actual <= inv.stock_minimo:
        log(db, "StockLow", f"Stock bajo {inv.product_name}: {inv.stock_actual}")
        await send_notification(authorization, f"ALERTA DE STOCK BAJO EN {inv.product_name} ({inv.branch_name}): {inv.stock_actual} UNIDADES")
    db.close(); return inv

@app.post("/inventory/output")
async def output_inventory(request: OutputRequest, authorization: str = Header(None), token=Depends(verify_token)):
    db = SessionLocal()
    inv = find_inv(db, request.product_id, request.branch_id)
    if not inv:
        db.close(); raise HTTPException(status_code=404, detail="Inventario no encontrado")
    if inv.stock_actual < request.quantity:
        db.close(); raise HTTPException(status_code=400, detail="Stock insuficiente")
    inv.stock_actual -= request.quantity
    add_kardex(db, inv, "SALIDA", request.quantity, request.reason)
    db.commit(); db.refresh(inv)
    log(db, "InventoryUpdated", f"Salida {request.quantity} producto {request.product_id}")
    if inv.stock_actual <= inv.stock_minimo:
        log(db, "StockLow", f"Stock bajo {inv.product_name}: {inv.stock_actual}")
        await send_notification(authorization, f"ALERTA DE STOCK BAJO EN {inv.product_name} ({inv.branch_name}): {inv.stock_actual} UNIDADES")
    db.close(); return inv

@app.post("/inventory/transfer")
async def transfer(request: TransferRequest, authorization: str = Header(None), token=Depends(verify_token)):
    db = SessionLocal()
    origin = find_inv(db, request.product_id, request.branch_origin_id)
    if not origin:
        db.close(); raise HTTPException(status_code=404, detail="Inventario origen no encontrado")
    if origin.stock_actual < request.quantity:
        db.close(); raise HTTPException(status_code=400, detail="Stock insuficiente")
    origin.stock_actual -= request.quantity
    add_kardex(db, origin, "TRANSFERENCIA_SALIDA", request.quantity, "Transferencia salida")
    dest = find_inv(db, request.product_id, request.branch_destination_id)
    if not dest:
        dest = Inventory(product_id=origin.product_id, product_code=origin.product_code, product_name=origin.product_name, company_name=origin.company_name, branch_id=request.branch_destination_id, branch_name=request.branch_destination_name, stock_actual=0, stock_minimo=origin.stock_minimo, cost=origin.cost, price=origin.price)
        db.add(dest); db.commit(); db.refresh(dest)
    dest.stock_actual += request.quantity
    add_kardex(db, dest, "TRANSFERENCIA_ENTRADA", request.quantity, "Transferencia entrada")
    db.commit(); db.refresh(origin); db.refresh(dest)
    log(db, "TransferCompleted", f"Transferencia {request.quantity} de {origin.branch_name} a {dest.branch_name}")
    if origin.stock_actual <= origin.stock_minimo:
        log(db, "StockLow", f"Stock bajo {origin.product_name}: {origin.stock_actual}")
        await send_notification(authorization, f"ALERTA DE STOCK BAJO EN {origin.product_name} ({origin.branch_name}): {origin.stock_actual} UNIDADES")
    db.close(); return {"message": "Transferencia completada", "origin_stock": origin.stock_actual, "destination_stock": dest.stock_actual}

@app.post("/inventory/loadExcel")
async def load_excel(file: UploadFile = File(...), token=Depends(verify_token)):
    db = SessionLocal()
    try:
        wb = openpyxl.load_workbook(file.file); sheet = wb.active; count = 0
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None: continue
            product_id, product_code, product_name, company_name, branch_id, branch_name, quantity, cost, price, stock_minimo = row
            inv = find_inv(db, int(product_id), int(branch_id))
            if not inv:
                inv = Inventory(product_id=int(product_id), product_code=str(product_code), product_name=str(product_name), company_name=str(company_name), branch_id=int(branch_id), branch_name=str(branch_name), stock_actual=0, stock_minimo=int(stock_minimo), cost=float(cost), price=float(price))
                db.add(inv); db.commit(); db.refresh(inv)
            inv.stock_actual += int(quantity)
            add_kardex(db, inv, "EXCEL", int(quantity), "Carga desde Excel")
            count += 1
        db.commit(); log(db, "InventoryLoaded", f"Excel importado: {count} filas"); db.close()
        return {"rows": count}
    except Exception as e:
        db.close(); raise HTTPException(status_code=400, detail=str(e))

@app.get("/inventory/product/{product_id}")
def by_product(product_id: int, token=Depends(verify_token)):
    db = SessionLocal(); data = db.query(Inventory).filter(Inventory.product_id == product_id).all(); db.close(); return data

@app.get("/inventory/product/{product_id}/consolidated")
def consolidated_by_product(product_id: int, token=Depends(verify_token)):
    db = SessionLocal()
    rows = db.query(Inventory).filter(Inventory.product_id == product_id).all()
    db.close()
    detalle = [
        {"empresa": r.company_name, "sucursal": r.branch_name, "branch_id": r.branch_id, "stock": r.stock_actual}
        for r in rows
    ]
    return {
        "product_id": product_id,
        "product_code": rows[0].product_code if rows else "",
        "product_name": rows[0].product_name if rows else "",
        "sucursales": len(rows),
        "saldo_total": sum(r.stock_actual for r in rows),
        "detalle": detalle,
    }

@app.get("/inventory/balance")
def balance(token=Depends(verify_token)):
    db = SessionLocal(); items = db.query(Inventory).all(); total = sum(i.stock_actual for i in items); db.close(); return {"total_general": total, "items": items}

@app.get("/inventory/kardex/{product_id}")
def get_kardex(product_id: int, token=Depends(verify_token)):
    db = SessionLocal(); data = db.query(Kardex).filter(Kardex.product_id == product_id).all(); db.close(); return data

@app.get("/inventory/events")
def events(token=Depends(verify_token)):
    db = SessionLocal(); data = db.query(EventLog).all(); db.close(); return data

@app.get("/reports/inventory.xlsx")
def export_inventory_excel(token=Depends(verify_token)):
    db = SessionLocal()
    items = db.query(Inventory).all()
    db.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "INVENTARIO"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = "A2:H2"

    title_fill = PatternFill("solid", fgColor="0B1F3A")
    header_fill = PatternFill("solid", fgColor="1D4ED8")
    low_fill = PatternFill("solid", fgColor="FEE2E2")
    ok_fill = PatternFill("solid", fgColor="DCFCE7")
    title_font = Font(color="FFFFFF", bold=True, size=14)
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells("A1:H1")
    ws["A1"] = "REPORTE PROFESIONAL DE INVENTARIO"
    ws["A1"].fill = title_fill
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    ws.append(["EMPRESA", "SUCURSAL", "PRODUCTO", "STOCK ACTUAL", "STOCK MINIMO", "COSTO", "PRECIO", "ALERTA"])
    for cell in ws[2]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    for item in items:
        alert = "STOCK BAJO" if item.stock_actual < 10 else "OK"
        ws.append([
            item.company_name,
            item.branch_name,
            item.product_name,
            item.stock_actual,
            item.stock_minimo,
            item.cost,
            item.price,
            alert,
        ])
        current_row = ws.max_row
        row_fill = low_fill if item.stock_actual < 10 else ok_fill
        for cell in ws[current_row]:
            cell.border = border
            cell.alignment = Alignment(vertical="center")
            if cell.column in (4, 5, 6, 7):
                cell.number_format = '#,##0.00'
            cell.fill = row_fill
        ws[f"H{current_row}"].font = Font(bold=True, color="991B1B" if item.stock_actual < 10 else "166534")

    total_row = ws.max_row + 2
    ws[f"A{total_row}"] = "TOTAL DE PRODUCTOS"
    ws[f"B{total_row}"] = len(items)
    ws[f"D{total_row}"] = "STOCK TOTAL"
    ws[f"E{total_row}"] = sum(item.stock_actual for item in items)
    ws[f"A{total_row}"].font = Font(bold=True)
    ws[f"D{total_row}"].font = Font(bold=True)
    ws[f"E{total_row}"].font = Font(bold=True)
    ws[f"E{total_row}"].number_format = '#,##0'

    for column, width in {"A": 24, "B": 24, "C": 34, "D": 14, "E": 16, "F": 14, "G": 14, "H": 16}.items():
        ws.column_dimensions[column].width = width

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    headers = {"Content-Disposition": 'attachment; filename="REPORTE_INVENTARIO.xlsx"'}
    return StreamingResponse(buffer, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers)
